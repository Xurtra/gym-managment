import argparse
import os
import shutil
import subprocess
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl


DONE_VALUES = {
    "yes",
    "y",
    "true",
    "done",
    "complete",
    "completed",
    "1",
    "x",
}


DEFAULT_KEEP_GOING = (
    "Continue with the next unfinished checklist item from the spreadsheet. "
    "Implement only that item, run the build command, and fix any errors."
)


def find_command(name):
    found = shutil.which(name)
    if found:
        return found

    if os.name == "nt":
        found = shutil.which(name + ".cmd")
        if found:
            return found

    return name


def run_process(cmd, cwd=None, input_text=None):
    printable = " ".join(f'"{x}"' if " " in str(x) else str(x) for x in cmd)
    print("\n> " + printable)

    result = subprocess.run(
        cmd,
        cwd=cwd,
        input=input_text,
        text=True,
        capture_output=True,
    )

    output = ""

    if result.stdout:
        print(result.stdout)
        output += result.stdout

    if result.stderr:
        print(result.stderr)
        output += "\n" + result.stderr

    return output, result.returncode


def run_shell(command, cwd=None):
    print("\n> " + command)

    result = subprocess.run(
        command,
        cwd=cwd,
        shell=True,
        text=True,
        capture_output=True,
    )

    output = ""

    if result.stdout:
        print(result.stdout)
        output += result.stdout

    if result.stderr:
        print(result.stderr)
        output += "\n" + result.stderr

    return output, result.returncode


def ensure_ollama_running(model_name):
    if ollama_is_running():
        print("\nOllama server is running.")
    else:
        print("\nOllama server is not running. Starting Ollama...")

        ollama_cmd = find_command("ollama")

        try:
            subprocess.Popen(
                [ollama_cmd, "serve"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except Exception as error:
            print(f"Could not start Ollama automatically: {error}")
            print("Open Ollama manually, then rerun this script.")
            sys.exit(1)

        for _ in range(10):
            time.sleep(1)
            if ollama_is_running():
                print("Ollama started.")
                break
        else:
            print("Ollama did not start in time. Open Ollama manually, then rerun.")
            sys.exit(1)

    ensure_ollama_model_installed(model_name)


def ollama_is_running():
    try:
        with urllib.request.urlopen("http://localhost:11434/api/tags", timeout=5):
            return True
    except Exception:
        return False


def ensure_ollama_model_installed(model_name):
    try:
        with urllib.request.urlopen("http://localhost:11434/api/tags", timeout=10) as response:
            data = response.read().decode("utf-8")
    except Exception:
        print("Could not read Ollama models.")
        return

    if model_name in data:
        print(f"Ollama model found: {model_name}")
        return

    print(f"\nOllama model missing: {model_name}")
    print("Pulling model now...")

    ollama_cmd = find_command("ollama")
    _, code = run_process([ollama_cmd, "pull", model_name])

    if code != 0:
        print(f"Failed to pull Ollama model: {model_name}")
        sys.exit(1)


def ask_ollama_keep_going(model_name, context):
    prompt = f"""
You are only a keep-going controller for Codex.

Your job:
Return one short instruction telling Codex what to do next.

Important rules:
- Never tell Codex to stop.
- Never say the work is complete.
- If the build failed, tell Codex to fix the build errors.
- If the build passed, tell Codex to move to the next unfinished checklist item.
- Keep the instruction direct and specific.
- Return only the instruction text.

Context:
{context}
"""

    payload = (
        "{"
        f'"model": "{model_name}", '
        f'"prompt": {json_escape(prompt)}, '
        '"stream": false'
        "}"
    ).encode("utf-8")

    request = urllib.request.Request(
        "http://localhost:11434/api/generate",
        data=payload,
        headers={"Content-Type": "application/json"},
    )

    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            raw = response.read().decode("utf-8")

        response_text = extract_ollama_response(raw)

        if not response_text:
            return DEFAULT_KEEP_GOING

        response_text = response_text.strip()

        if "stop" in response_text.lower() and len(response_text) < 120:
            return DEFAULT_KEEP_GOING

        return response_text

    except Exception as error:
        print(f"Ollama keep-going failed: {error}")
        return DEFAULT_KEEP_GOING


def json_escape(text):
    import json
    return json.dumps(text)


def extract_ollama_response(raw):
    import json

    try:
        data = json.loads(raw)
        return str(data.get("response", "")).strip()
    except Exception:
        return ""


def ensure_node_project_ready(project_dir):
    package_json = project_dir / "package.json"

    if not package_json.exists():
        print("\nNo package.json found. Skipping npm setup.")
        return

    print("\nChecking Node/TypeScript setup...")

    if not (project_dir / "node_modules").exists():
        print("\nnode_modules missing. Running npm install...")
        run_shell("npm install", cwd=project_dir)

    _, ts_code = run_shell("npm ls typescript --depth=0", cwd=project_dir)

    if ts_code != 0:
        print("\nTypeScript missing. Installing TypeScript...")
        run_shell("npm install -D typescript", cwd=project_dir)


def fix_missing_tsc_if_needed(test_output, project_dir):
    lowered = test_output.lower()

    missing_tsc_messages = [
        "tsc' is not recognized",
        '"tsc" is not recognized',
        "tsc is not recognized",
        "cannot find module 'typescript'",
        "cannot find package 'typescript'",
    ]

    if any(message in lowered for message in missing_tsc_messages):
        print("\nDetected missing TypeScript compiler. Installing TypeScript...")
        run_shell("npm install -D typescript", cwd=project_dir)
        return True

    return False


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def load_sheet(checklist_path):
    wb = openpyxl.load_workbook(checklist_path)
    ws = wb.active

    headers = [normalize(cell.value) for cell in ws[1]]

    completed_col = find_or_create_column(ws, headers, "Completed")
    ongoing_col = find_optional_column(headers, "Ongoing")
    taken_col = find_optional_column(headers, "Taken")

    wb.save(checklist_path)

    return wb, ws, headers, completed_col, ongoing_col, taken_col


def find_optional_column(headers, name):
    target = name.lower()

    for index, header in enumerate(headers, start=1):
        if header.lower() == target:
            return index

    return None


def find_or_create_column(ws, headers, name):
    target = name.lower()

    for index, header in enumerate(headers, start=1):
        if header.lower() == target:
            return index

    new_col = len(headers) + 1
    ws.cell(row=1, column=new_col).value = name
    headers.append(name)
    return new_col


def is_done(value):
    return normalize(value).lower() in DONE_VALUES


def row_is_empty(ws, row_index):
    for col_index in range(1, ws.max_column + 1):
        value = ws.cell(row=row_index, column=col_index).value
        if value not in [None, ""]:
            return False

    return True


def row_to_text(ws, headers, row_index):
    parts = []

    for col_index, header in enumerate(headers, start=1):
        if not header:
            continue

        value = ws.cell(row=row_index, column=col_index).value

        if value not in [None, ""]:
            parts.append(f"{header}: {value}")

    return " | ".join(parts)


def get_next_unfinished_row(checklist_path):
    wb, ws, headers, completed_col, ongoing_col, taken_col = load_sheet(checklist_path)

    for row_index in range(2, ws.max_row + 1):
        if row_is_empty(ws, row_index):
            continue

        completed_value = ws.cell(row=row_index, column=completed_col).value

        if not is_done(completed_value):
            row_text = row_to_text(ws, headers, row_index)
            return row_index, row_text

    return None, None


def mark_in_progress(checklist_path, row_index):
    wb, ws, headers, completed_col, ongoing_col, taken_col = load_sheet(checklist_path)

    if ongoing_col:
        ws.cell(row=row_index, column=ongoing_col).value = "YES"

    if taken_col:
        ws.cell(row=row_index, column=taken_col).value = "YES"

    wb.save(checklist_path)


def mark_completed(checklist_path, row_index):
    wb, ws, headers, completed_col, ongoing_col, taken_col = load_sheet(checklist_path)

    ws.cell(row=row_index, column=completed_col).value = "YES"

    if ongoing_col:
        ws.cell(row=row_index, column=ongoing_col).value = ""

    if taken_col:
        ws.cell(row=row_index, column=taken_col).value = "YES"

    wb.save(checklist_path)


def build_codex_prompt(task, row_index, row_text, test_command, keep_going_instruction, previous_error):
    error_section = ""

    if previous_error:
        error_section = f"""
Previous build/test failed.

Error output:
{previous_error[-5000:]}

Fix those errors before doing anything else.
"""

    return f"""
You are working inside this TypeScript project.

Overall goal:
{task}

Current checklist row:
Row {row_index}: {row_text}

Keep-going instruction from Ollama:
{keep_going_instruction}

Rules:
1. Implement ONLY the current checklist row.
2. Do not skip to another checklist row.
3. Do not invent unrelated features.
4. Do not edit unrelated files unless needed for this row.
5. Do not edit the spreadsheet. The automation script will update it.
6. Run this command after implementing:
   {test_command}
7. Fix all TypeScript/build errors from that command.
8. Stop after this row is implemented and the build passes.

{error_section}
"""


def build_codex_command(args, first_run):
    codex_cmd = find_command("codex")

    if first_run:
        cmd = [
            codex_cmd,
            "exec",
            "--sandbox",
            "workspace-write",
            "--ask-for-approval",
            "never",
        ]

        if args.codex_model:
            cmd += ["--model", args.codex_model]

        if args.reasoning:
            cmd += ["--config", f'model_reasoning_effort="{args.reasoning}"']

        if args.skip_git_repo_check:
            cmd += ["--skip-git-repo-check"]

        cmd += ["-"]
        return cmd

    cmd = [
        codex_cmd,
        "exec",
        "resume",
        "--last",
        "-",
    ]

    return cmd


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--checklist", required=True)
    parser.add_argument("--task", required=True)
    parser.add_argument("--test", default="npm run build")
    parser.add_argument("--ollama-model", default="qwen2.5-coder:7b")
    parser.add_argument("--codex-model", default="")
    parser.add_argument("--reasoning", default="")
    parser.add_argument("--max-items", type=int, default=9999)
    parser.add_argument("--max-attempts-per-item", type=int, default=5)
    parser.add_argument("--skip-git-repo-check", action="store_true")
    args = parser.parse_args()

    project_dir = Path.cwd()
    checklist_path = Path(args.checklist).resolve()

    if not checklist_path.exists():
        print(f"Checklist file not found: {checklist_path}")
        sys.exit(1)

    ensure_ollama_running(args.ollama_model)
    ensure_node_project_ready(project_dir)

    completed_count = 0
    first_codex_run = True
    keep_going_instruction = DEFAULT_KEEP_GOING

    while completed_count < args.max_items:
        row_index, row_text = get_next_unfinished_row(checklist_path)

        if row_index is None:
            print("\nAll checklist entries are completed.")
            break

        print(f"\n================ CHECKLIST ROW {row_index} ================")
        print(row_text)

        mark_in_progress(checklist_path, row_index)

        previous_error = ""

        for attempt in range(1, args.max_attempts_per_item + 1):
            print(f"\nAttempt {attempt}/{args.max_attempts_per_item} for row {row_index}")

            prompt = build_codex_prompt(
                task=args.task,
                row_index=row_index,
                row_text=row_text,
                test_command=args.test,
                keep_going_instruction=keep_going_instruction,
                previous_error=previous_error,
            )

            codex_cmd = build_codex_command(args, first_codex_run)

            codex_output, codex_code = run_process(
                codex_cmd,
                cwd=project_dir,
                input_text=prompt,
            )

            first_codex_run = False

            test_output, test_code = run_shell(args.test, cwd=project_dir)

            if fix_missing_tsc_if_needed(test_output, project_dir):
                test_output, test_code = run_shell(args.test, cwd=project_dir)

            context = f"""
Checklist row:
Row {row_index}: {row_text}

Codex exit code:
{codex_code}

Latest Codex output:
{codex_output[-4000:]}

Build command:
{args.test}

Build exit code:
{test_code}

Build output:
{test_output[-4000:]}
"""

            keep_going_instruction = ask_ollama_keep_going(
                args.ollama_model,
                context,
            )

            print("\nOllama keep-going instruction:")
            print(keep_going_instruction)

            if test_code == 0:
                print(f"\nBuild passed. Marking row {row_index} completed.")
                mark_completed(checklist_path, row_index)
                completed_count += 1
                break

            print(f"\nBuild failed for row {row_index}. Codex will retry.")
            previous_error = test_output

        else:
            print(f"\nRow {row_index} failed after max attempts.")
            print("Stopping so it does not loop forever on the same item.")
            print("Increase --max-attempts-per-item if you want it to keep trying.")
            break

    print(f"\nFinished. Completed {completed_count} checklist item(s) this run.")


if __name__ == "__main__":
    main()