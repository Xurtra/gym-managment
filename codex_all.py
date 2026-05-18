import argparse
import os
import subprocess
from pathlib import Path

import openpyxl


DONE_VALUES = {"yes", "y", "true", "done", "complete", "completed", "1", "x"}


def run_command(args, cwd=None, shell_string=False):
    if shell_string:
        cmd = args
        printable = args
        use_shell = True
    else:
        printable = " ".join(args)
        if os.name == "nt":
            cmd = subprocess.list2cmdline(args)
            use_shell = True
        else:
            cmd = args
            use_shell = False

    print("\n> " + printable)

    result = subprocess.run(
        cmd,
        cwd=cwd,
        shell=use_shell,
        text=True,
        capture_output=True,
    )

    output = ""

    if result.stdout:
        print(result.stdout)
        output += result.stdout

    if result.stderr:
        print(result.stderr)
        output += "\n" + result.stderr

    return output, result.returncode


def ensure_node_ready(project_dir):
    package_json = project_dir / "package.json"

    if not package_json.exists():
        print("No package.json found. Skipping npm setup.")
        return

    if not (project_dir / "node_modules").exists():
        print("node_modules missing. Running npm install...")
        run_command("npm install", cwd=project_dir, shell_string=True)


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def find_or_create_column(ws, headers, name):
    lowered = [h.lower() for h in headers]

    if name.lower() in lowered:
        return lowered.index(name.lower()) + 1

    col = len(headers) + 1
    ws.cell(row=1, column=col).value = name
    headers.append(name)
    return col


def load_sheet(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(normalize(cell.value))

    completed_col = find_or_create_column(ws, headers, "Completed")

    ongoing_col = None
    taken_col = None

    for i, header in enumerate(headers, start=1):
        h = header.lower()
        if h == "ongoing":
            ongoing_col = i
        if h == "taken":
            taken_col = i

    return wb, ws, headers, completed_col, ongoing_col, taken_col


def is_done(value):
    return normalize(value).lower() in DONE_VALUES


def row_to_text(ws, headers, row_index):
    parts = []

    for col_index, header in enumerate(headers, start=1):
        if not header:
            continue

        value = ws.cell(row=row_index, column=col_index).value

        if value not in [None, ""]:
            parts.append(f"{header}: {value}")

    return " | ".join(parts)


def get_next_unfinished(path):
    wb, ws, headers, completed_col, ongoing_col, taken_col = load_sheet(path)

    for row_index in range(2, ws.max_row + 1):
        row_text = row_to_text(ws, headers, row_index)

        if not row_text.strip():
            continue

        completed_value = ws.cell(row=row_index, column=completed_col).value

        if not is_done(completed_value):
            return wb, ws, headers, row_index, completed_col, ongoing_col, taken_col, row_text

    return wb, ws, headers, None, completed_col, ongoing_col, taken_col, None


def mark_in_progress(path, row_index):
    wb, ws, headers, completed_col, ongoing_col, taken_col = load_sheet(path)

    if ongoing_col:
        ws.cell(row=row_index, column=ongoing_col).value = "YES"

    if taken_col:
        ws.cell(row=row_index, column=taken_col).value = "YES"

    wb.save(path)


def mark_completed(path, row_index):
    wb, ws, headers, completed_col, ongoing_col, taken_col = load_sheet(path)

    ws.cell(row=row_index, column=completed_col).value = "YES"

    if ongoing_col:
        ws.cell(row=row_index, column=ongoing_col).value = ""

    if taken_col:
        ws.cell(row=row_index, column=taken_col).value = "YES"

    wb.save(path)


def fix_missing_tsc_if_needed(output, project_dir):
    lowered = output.lower()

    if "tsc' is not recognized" in lowered or "tsc is not recognized" in lowered:
        print("\nTypeScript compiler missing. Installing TypeScript...")
        run_command("npm install -D typescript", cwd=project_dir, shell_string=True)
        return True

    return False


def make_prompt(task, row_index, row_text, test_command, error_output=""):
    error_section = ""

    if error_output:
        error_section = f"""
Previous build/test failed.

Error output:
{error_output[-5000:]}

Fix these errors before marking the item complete.
"""

    return f"""
You are working inside this TypeScript project.

Overall task:
{task}

Current checklist row:
Row {row_index}: {row_text}

Instructions:
1. Implement ONLY this checklist row.
2. Do not skip to another row.
3. Do not invent unrelated features.
4. Update code as needed.
5. Run this command:
   {test_command}
6. Fix all TypeScript/build errors.
7. When the row is actually implemented and the build passes, stop.

{error_section}
"""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--checklist", required=True)
    parser.add_argument("--task", required=True)
    parser.add_argument("--test", default="npm run build")
    parser.add_argument("--max-items", type=int, default=9999)
    parser.add_argument("--max-attempts-per-item", type=int, default=5)
    args = parser.parse_args()

    project_dir = Path.cwd()
    checklist_path = Path(args.checklist).resolve()

    ensure_node_ready(project_dir)

    first_codex_run = True
    completed_count = 0

    while completed_count < args.max_items:
        result = get_next_unfinished(checklist_path)
        wb, ws, headers, row_index, completed_col, ongoing_col, taken_col, row_text = result

        if row_index is None:
            print("\nAll checklist entries are completed.")
            break

        print(f"\n================ CHECKLIST ROW {row_index} ================")
        print(row_text)

        mark_in_progress(checklist_path, row_index)

        last_error = ""

        for attempt in range(1, args.max_attempts_per_item + 1):
            print(f"\nAttempt {attempt}/{args.max_attempts_per_item} for row {row_index}")

            prompt = make_prompt(
                args.task,
                row_index,
                row_text,
                args.test,
                last_error
            )

            if first_codex_run:
                codex_cmd = [
                    "codex",
                    "exec",
                    "--sandbox",
                    "workspace-write",
                    "--ask-for-approval",
                    "never",
                    prompt,
                ]
                first_codex_run = False
            else:
                codex_cmd = [
                    "codex",
                    "exec",
                    "resume",
                    "--last",
                    prompt,
                ]

            run_command(codex_cmd, cwd=project_dir)

            test_output, test_code = run_command(
                args.test,
                cwd=project_dir,
                shell_string=True
            )

            if fix_missing_tsc_if_needed(test_output, project_dir):
                test_output, test_code = run_command(
                    args.test,
                    cwd=project_dir,
                    shell_string=True
                )

            if test_code == 0:
                print(f"\nBuild passed. Marking row {row_index} completed.")
                mark_completed(checklist_path, row_index)
                completed_count += 1
                break

            print(f"\nBuild failed for row {row_index}. Codex will retry.")
            last_error = test_output

        else:
            print(f"\nRow {row_index} failed after max attempts. Stopping.")
            print("Fix this row manually or increase --max-attempts-per-item.")
            break

    print(f"\nFinished. Completed {completed_count} checklist item(s) this run.")


if __name__ == "__main__":
    main()